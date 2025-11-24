from fastapi import FastAPI, Form
from fastapi.responses import HTMLResponse
from openpyxl import Workbook, load_workbook
import os

## \file main2.py
## \brief 论文提交与 Excel 存储服务
## \author 作业提交者
## \date 2025-11-23

app = FastAPI()

EXCEL_FILE = "homework1/work2/papers.xlsx"

## \brief 初始化 Excel 文件，若不存在则创建
## \return None

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["标题", "作者", "章节"])
        wb.save(EXCEL_FILE)

## \brief 写入一行论文信息到 Excel
## \param title 论文标题
## \param author 作者
## \param chapter 章节
## \return None

def save_to_excel(title, author, chapter):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([title, author, chapter])
    wb.save(EXCEL_FILE)

## \brief 首页，返回论文提交 HTML 页面
## \return HTMLResponse
@app.get("/", response_class=HTMLResponse)
async def home():
    return """
    <html>
        <head>
            <meta charset="utf-8">
            <title>论文提交页面</title>
        </head>
        <body>
            <h2>请输入论文信息</h2>
            <form action="/submit" method="post">
                <label>论文标题：</label><br>
                <input type="text" name="title" required><br><br>

                <label>作者：</label><br>
                <input type="text" name="author" required><br><br>

                <label>章节：</label><br>
                <input type="text" name="chapter" required><br><br>

                <button type="submit">提交</button>
            </form>
        </body>
    </html>
    """

## \brief 论文提交接口，保存数据到 Excel
## \param title 论文标题
## \param author 作者
## \param chapter 章节
## \return 提交结果字典
@app.post("/submit")
async def submit(
    title: str = Form(...),
    author: str = Form(...),
    chapter: str = Form(...)
):
    save_to_excel(title, author, chapter)
    return {"message": "提交成功！", "data": {"标题": title, "作者": author, "章节": chapter}}

## \brief 启动前初始化 Excel 文件
init_excel()
# 运行命令：

#uvicorn homework1.work2.main2:app --reload